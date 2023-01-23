import sys
from openpyxl import load_workbook
from jinja2 import Template

XML_TEMPLATE = '''<propertiesmap>
    {% for exclusion in exclusions %}<key name="{{exclusion.key}}">
        <key name="Hosts"></key>
        <key name="Ports"></key>
        <tBOOL name="bEnabled">1</tBOOL>
        <tDWORD name="AppID">0</tDWORD>
        <tSTRING name="sImagePath">{{exclusion.path}}</tSTRING>
        <tDWORD name="nHost_">0</tDWORD>
        <tWORD name="nPort_">0</tWORD>
        <tDWORD name="nTriggers">305</tDWORD>
        <tQWORD name="Hash">0</tQWORD>
        <tBOOL name="SSLOnly">0</tBOOL>
        <tSTRING name="sComment">{{exclusion.comment}}</tSTRING>
    </key>{% endfor %}
    <tDWORD name="unique_id">1475100830</tDWORD>
    <tSTRING name="SettingsVersion">21.4.20.669</tSTRING>
</propertiesmap>'''

HELP_TEXT = '''Usage:
    %s <xlsx file path> <output xml file path>''' %sys.argv[0]

def get_first_worksheet_from_xlsx(xlsx_path):
    try:
        wb = load_workbook(xlsx_path)
    except:
        print ("Incorrect <xlsx file path>")
        sys.exit(1)

    return wb[wb.sheetnames[0]]


def get_exclusions_from_worksheet(ws):
    exclusions = []
    key = 0
    for row in ws.iter_rows(min_row=2):
        exclusions.append(
            {
                "key": str(key).zfill(4),
                "path": row[0].value,
                "comment": row[1].value
            })
        key += 1

    return exclusions


def generate_xml_from_template(xml_template, exclusions):
    template = Template(xml_template)

    return template.render(exclusions=exclusions)


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print(HELP_TEXT)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    xml_path = sys.argv[2]

    ws = get_first_worksheet_from_xlsx(xlsx_path)
    exclusions = get_exclusions_from_worksheet(ws)
    xml_content = generate_xml_from_template (XML_TEMPLATE, exclusions)

    try:
        xml_file = open(xml_path, 'w')
        xml_file.write(xml_content)
    except:
        print("Error writing a file to the path %s" %xml_path)
        sys.exit(1)